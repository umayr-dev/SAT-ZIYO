#!/usr/bin/env python3
"""
Read Math scoring Excel file and generate Math 2D table
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

def read_math_excel(file_path):
    """Read Math Excel file and extract 2D table"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Get the first sheet (or specify sheet name if needed)
    sheet = wb.active
    print(f"Reading sheet: {sheet.title}")
    
    # Find the data range
    # Usually M1 is in rows, M2 is in columns
    # First row might be headers (M2 values)
    # First column might be M1 values
    
    table = {}
    
    # Find header row (usually row 1 or 2)
    header_row = None
    for row_idx in range(1, min(30, sheet.max_row + 1)):
        first_cell = sheet.cell(row=row_idx, column=1).value
        if first_cell is not None and (isinstance(first_cell, (int, float)) or str(first_cell).strip() in ['M1', 'Module 1', '']):
            header_row = row_idx
            break
    
    if header_row is None:
        header_row = 1
    
    print(f"Header row: {header_row}")
    
    # Read M2 values from header row (columns 2 onwards)
    m2_values = []
    for col_idx in range(2, min(30, sheet.max_column + 1)):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value is not None:
            try:
                m2_val = int(float(cell_value))
                m2_values.append((col_idx, m2_val))
            except (ValueError, TypeError):
                pass
    
    print(f"M2 values found: {[v[1] for v in m2_values]}")
    
    # Read M1 values and scores
    data_start_row = header_row + 1
    for row_idx in range(data_start_row, min(50, sheet.max_row + 1)):
        m1_cell = sheet.cell(row=row_idx, column=1).value
        if m1_cell is None:
            continue
        
        try:
            m1_val = int(float(m1_cell))
        except (ValueError, TypeError):
            continue
        
        m1_row = {}
        for col_idx, m2_val in m2_values:
            score_cell = sheet.cell(row=row_idx, column=col_idx).value
            if score_cell is not None:
                try:
                    score = int(float(score_cell))
                    m1_row[m2_val] = score
                except (ValueError, TypeError):
                    pass
        
        if m1_row:
            table[m1_val] = m1_row
    
    return table

if __name__ == "__main__":
    excel_file = "public/scoring math.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found")
        sys.exit(1)
    
    print(f"Reading {excel_file}...")
    table = read_math_excel(excel_file)
    
    # Print as JavaScript object
    print("\n// Math 2D Table from Excel:")
    print("const digitalSATMath2DTable = {")
    
    for m1 in sorted(table.keys(), reverse=True):
        print(f"  {m1}: {{")
        for m2 in sorted(table[m1].keys(), reverse=True):
            score = table[m1][m2]
            print(f"    {m2}: {score},")
        print("  },")
    
    print("};")
    
    # Also save as JSON for reference
    with open("math_table.json", "w") as f:
        json.dump(table, f, indent=2)
    
    print(f"\nTable saved to math_table.json")
    print(f"Total M1 values: {len(table)}")

